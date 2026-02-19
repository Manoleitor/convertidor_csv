import unittest
from pathlib import Path
from process_signatures import (
    get_participants,
    create_sign_sheet_workbook,
    save_workbook,
    create_sign_pdf_html,
    save_pdf_from_html,
    _setup_styles,
    _get_cell_border,
    _setup_column_widths,
    _write_participant_tables,
    _write_headers,
    _write_single_table,
    create_sign_sheet_workbook
)
import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Alignment
from openpyxl import Workbook

class TestProcessSignatures(unittest.TestCase):
    def setUp(self):
        # 53 nombres de Pokémon (sin repeticiones excesivas)
        pokemon_names = [
            "Pikachu", "Bulbasaur", "Charmander", "Squirtle", "Jigglypuff", "Meowth", "Psyduck", "Machop", "Magnemite", "Eevee",
            "Snorlax", "Mewtwo", "Chikorita", "Cyndaquil", "Totodile", "Mareep", "Sudowoodo", "Espeon", "Umbreon", "Murkrow",
            "Wobbuffet", "Larvitar", "Treecko", "Torchic", "Mudkip", "Ralts", "Surskit", "Shroomish", "Makuhita", "Skitty",
            "Sableye", "Mawile", "Meditite", "Swablu", "Barboach", "Bagon", "Beldum", "Turtwig", "Chimchar", "Piplup",
            "Starly", "Kricketot", "Shinx", "Riolu", "Gible", "Hippopotas", "Snover", "Rotom", "Snivy", "Tepig",
            "Oshawott", "Zorua", "Axew"
        ]
        french_surnames = [
            "Dufour", "Lafitte", "Dumont", "Duchamp", "Bourdet", "Bideau", "Lissarrague", "Darrigrand", "Larrieu", "Larronde",
            "Larrarte", "Larrucea", "Etcheto", "Etcheverry", "Etchegoyen", "Etcheberry", "Larralde", "Darracq", "Bidegain", "Larranaga",
            "Lafargue", "Laborde", "Lafon", "Lafargue", "Lafitte", "Lafon", "Lafargue", "Lafitte", "Lafon", "Lafargue"
        ]
        basque_surnames = [
            "Aguirre", "Echeverria", "Goikoetxea", "Ibarra", "Irizar", "Mendieta", "Oteiza", "Urrutia", "Zabala", "Zubizarreta",
            "Arrieta", "Etxeberria", "Garate", "Garmendia", "Goñi", "Iturriaga", "Lasa", "Muguruza", "Olaizola", "Sarasola",
            "Ugalde", "Urkiza", "Zuloaga", "Zunzunegui", "Zubiri", "Zubia", "Zubeldia", "Zubimendi", "Zubizarreta", "Zuloaga", "Zubia"
        ]
        self.test_participants = [
            (pokemon_names[i], f"{french_surnames[i % len(french_surnames)]} {basque_surnames[i % len(basque_surnames)]}") for i in range(53)
        ]
        self.result_dir = Path("./result_test")
        self.result_dir.mkdir(exist_ok=True)
        self.excel_path = self.result_dir / "test_hoja_firmas.xlsx"
        self.pdf_path = self.result_dir / "test_hoja_firmas.pdf"

    def tearDown(self):
        if self.excel_path.exists():
            self.excel_path.unlink()
        if self.pdf_path.exists():
            self.pdf_path.unlink()
        if self.result_dir.exists():
            try:
                self.result_dir.rmdir()
            except Exception:
                pass

    def test_create_sign_sheet_workbook_and_save(self):
        wb = create_sign_sheet_workbook(self.test_participants)
        save_workbook(wb, self.excel_path)
        self.assertTrue(self.excel_path.exists())
        # Comprobar que la hoja tiene el nombre correcto
        loaded = load_workbook(self.excel_path)
        self.assertIn("Firmas", loaded.sheetnames)

    def test_setup_styles(self):
        styles = _setup_styles()
        self.assertIn('thin_border', styles)
        self.assertIn('thick_border', styles)
        self.assertIn('left_alignment', styles)
        self.assertIn('center_alignment', styles)
        self.assertIsInstance(styles['thin_border'], Border)
        self.assertIsInstance(styles['thick_border'], Border)
        self.assertIsInstance(styles['left_alignment'], Alignment)
        self.assertIsInstance(styles['center_alignment'], Alignment)

    def test_setup_column_widths(self):
        wb = Workbook()
        ws = wb.active
        _setup_column_widths(ws)
        # Check that the first three columns have the expected width
        self.assertGreaterEqual(ws.column_dimensions['A'].width, 10)
        self.assertGreaterEqual(ws.column_dimensions['B'].width, 10)
        self.assertGreaterEqual(ws.column_dimensions['C'].width, 10)
        self.assertGreaterEqual(ws.column_dimensions['D'].width, 5)
        self.assertGreaterEqual(ws.column_dimensions['E'].width, 5)
        self.assertGreaterEqual(ws.column_dimensions['F'].width, 10)
        self.assertGreaterEqual(ws.column_dimensions['G'].width, 10)
        self.assertGreaterEqual(ws.column_dimensions['H'].width, 10)

    #test para _write_headers
    def test_write_headers(self):
        wb = Workbook()
        ws = wb.active
        styles = _setup_styles()
        _write_headers(ws, styles)
        self.assertEqual(ws.cell(row=1, column=1).value, "Nombre")
        self.assertEqual(ws.cell(row=1, column=2).value, "Apellidos")
        # self.assertEqual(ws.cell(row=1, column=6).value, "Nombre")
        # self.assertEqual(ws.cell(row=1, column=7).value, "Apellidos")

    def test_create_sign_pdf_html_and_save(self):
        html = create_sign_pdf_html(self.test_participants, subject="Test", group="T1", week=1, day1="Lunes", day2="Viernes")
        self.assertIn("Hoja de Firmas", html)
        save_pdf_from_html(html, self.pdf_path)
        self.assertTrue(self.pdf_path.exists())

    def test_create_sign_pdf_html_placeholders(self):
        html = create_sign_pdf_html(self.test_participants)
        self.assertIn("primera", html)
        self.assertIn("segunda", html)

    def test_create_sign_sheet_workbook_empty(self):
        wb = create_sign_sheet_workbook([])
        save_workbook(wb, self.excel_path)
        self.assertTrue(self.excel_path.exists())

    def test_create_sign_pdf_html_special_characters(self):
        # Test PDF con caracteres especiales en nombres
        participants = [("José María", "Muñoz-Álvarez"), ("Zoë", "O'Connor")]
        html = create_sign_pdf_html(participants, subject="Español & Français", group="T2", week=2, day1="Lunes", day2="Miércoles")
        self.assertIn("Español & Français", html)
        save_pdf_from_html(html, self.pdf_path)
        self.assertTrue(self.pdf_path.exists())

    def test_create_sign_pdf_html_optional_args(self):
        html = create_sign_pdf_html(self.test_participants)
        self.assertIn("Hoja de Firmas", html)
        save_pdf_from_html(html, self.pdf_path)
        self.assertTrue(self.pdf_path.exists())

    def test_get_participants_empty(self):
        self.assertIsInstance(get_participants(), list)

if __name__ == "__main__":
    unittest.main()
