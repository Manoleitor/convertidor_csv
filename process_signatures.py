"""Script para procesar archivo CSV y generar hoja de firmas.

Lee un archivo CSV de la carpeta Downloads, extrae nombres y apellidos,
y genera una hoja de cálculo con participantes agrupados de 8 en 8
organizados en 3 columnas por bloque.
"""


import re
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
# Nuevas importaciones para PDF
from weasyprint import HTML, CSS
from jinja2 import Template


def find_participants_file() -> Path:
    """Busca el primer archivo que coincida con el patrón courseid_XXX_participants.

    Returns:
        Path al archivo encontrado, o None si no hay coincidencias.
    """
    downloads_dir = Path.home() / "Downloads"
    pattern = r"courseid_\d{3}_participants\.csv"

    for file in sorted(downloads_dir.glob("courseid_*_participants.csv")):
        if re.match(pattern, file.name):
            return file

    return None



def get_participants() -> list:
    """
    Obtiene la lista de participantes ordenados alfabéticamente.

    Returns:
        Lista de participantes (str).
    """
    csv_path = find_participants_file()
    if not csv_path:
        print("Error: No se encontró ningún archivo con patrón courseid_XXX_participants.csv en Downloads")
        return []
    print(f"✓ Archivo encontrado: {csv_path.name}")
    try:
        df = pd.read_csv(csv_path)
    except Exception as e:
        print(f"Error al leer el archivo CSV: {e}")
        return []
    required_columns = ["Nombre", "Apellido(s)"]
    if not all(col in df.columns for col in required_columns):
        print(f"Error: El archivo no contiene las columnas {required_columns}")
        print(f"Columnas encontradas: {list(df.columns)}")
        return []
    df["Nombre"] = df["Nombre"].str.strip()
    df["Apellido(s)"] = df["Apellido(s)"].str.strip()
    participants = list(zip(df["Nombre"], df["Apellido(s)"]))
    return sorted(participants, key=lambda x: x[0].lower())

def create_sign_sheet(participant_list: list, output_path: Path) -> None:
    """
    Crea y guarda la hoja de firmas en Excel.

    Args:
        participant_list: Lista de participantes (str).
        output_path: Ruta de salida del archivo Excel.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Firmas"
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    thick_border = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick"),
    )
    left_alignment = Alignment(horizontal="left", vertical="center")
    center_alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 5
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 12

    ws.row_dimensions[1].height = 20
    headers = ["Nombre", "Apellidos", "Firma"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.font = Font(bold=True)
    # Nueva lógica: escribir tablas de 8 participantes en parejas horizontales
    tables_per_row = 2  # número de tablas por fila
    table_rows = 8      # filas por tabla
    table_cols = 3      # columnas por tabla
    table_h_spacing = 2 # espacio entre tablas horizontal
    table_v_spacing = 2 # espacio entre tablas vertical

    num_tables = (len(participant_list) + table_rows - 1) // table_rows
    for table_idx in range(num_tables):
        # Calcular posición de la tabla
        row_block = table_idx // tables_per_row
        col_block = table_idx % tables_per_row
        start_row = 2 + row_block * (table_rows + table_v_spacing)
        start_col = 1 + col_block * (table_cols + table_h_spacing)

        # Obtener participantes de este bloque
        block_participants = participant_list[table_idx * table_rows : (table_idx + 1) * table_rows]

        for row_in_table, participant in enumerate(block_participants):
            first_name, last_name = participant
            is_first_row = row_in_table == 0
            is_last_row = row_in_table == table_rows - 1 or row_in_table == len(block_participants) - 1
            for col_in_table in range(table_cols):
                col = start_col + col_in_table
                left = "thick" if col_in_table == 0 else "thin"
                right = "thick" if col_in_table == table_cols - 1 else "thin"
                top = "thick" if is_first_row else "thin"
                bottom = "thick" if is_last_row else "thin"
                border = Border(
                    left=Side(style=left),
                    right=Side(style=right),
                    top=Side(style=top),
                    bottom=Side(style=bottom),
                )
                if col_in_table == 0:
                    value = first_name
                    alignment = left_alignment
                elif col_in_table == 1:
                    value = last_name
                    alignment = left_alignment
                else:
                    value = ""
                    alignment = left_alignment
                cell = ws.cell(row=start_row + row_in_table, column=col, value=value)
                cell.border = border
                cell.alignment = alignment
            ws.row_dimensions[start_row + row_in_table].height = 25
    try:
        wb.save(output_path)
        print(f"✓ Archivo generado exitosamente en: {output_path}")
        print(f"✓ Total de participantes procesados: {len(participant_list)}")
    except Exception as e:
        print(f"Error al guardar el archivo: {e}")

# Nuevo método para crear PDF de hoja de firmas
def create_sign_pdf(participant_list: list, output_path: Path, subject: str = None, group: str = None, week: int = None, day1: str = 'primera', day2: str = 'segunda') -> None:
    """
    Crea y guarda la hoja de firmas en PDF usando WeasyPrint y Jinja2.

    Args:
        participant_list: Lista de participantes (str).
        output_path: Ruta de salida del archivo PDF.
        subject: Nombre de la asignatura (opcional).
    """
    # Agrupar participantes en bloques de 8
    table_rows = 8
    tables_per_row = 2
    blocks = [participant_list[i:i+table_rows] for i in range(0, len(participant_list), table_rows)]
    # Agrupar bloques en filas de dos
    block_pairs = [blocks[i:i+tables_per_row] for i in range(0, len(blocks), tables_per_row)]

    # HTML y CSS para la hoja de firmas
    html_template = """
    <!DOCTYPE html>
    <html lang='es'>
    <head>
        <meta charset='UTF-8'>
        <title>Hoja de Firmas</title>
        <style>
        @page { size: A4; margin: 10mm 20mm 20mm 10mm; }
        body {
            font-family: 'Segoe UI', 'Arial', sans-serif;
            margin: 0;
            color: #222;
        }
        .header-container {
            display: flex;
            align-items: center;
            gap: 18px;
            margin: 20px 0 30px 0;
        }
        .subject-box, .group-box, .date-box, .week-box {
            background: #f5f5f5;
            border: 1px solid #888;
            border-radius: 6px;
            padding: 6px 18px;
            font-size: 13px;
            font-weight: bold;
            color: #222;
            letter-spacing: 1px;
            box-shadow: 0 1px 4px #bbb;
        }
        .subject-box { background: {{ '#e0e0e0' if subject else '#f5f5f5' }}; }
        .group-box { background: {{ '#e0e0e0' if group else '#f5f5f5' }}; }
        .week-box { background: {{ '#e0e0e0' if week else '#f5f5f5' }}; }
        .subject-placeholder, .group-placeholder, .date-placeholder, .week-placeholder {
            color: #aaa;
        }
        h2 {
            font-size: 18px;
            color: #222;
            letter-spacing: 1px;
            margin: 0;
        }
        .subject-box {
            background: {{ '#e0e0e0' if subject else '#f5f5f5' }};
            border: 1px solid #888;
            border-radius: 6px;
            padding: 6px 18px;
            font-size: 13px;
            font-weight: bold;
            color: #222;
            letter-spacing: 1px;
            box-shadow: 0 1px 4px #bbb;
        }
        .subject-placeholder {
            color: #aaa;
        }
        .row {
            display: flex;
            flex-direction: row;
            justify-content: flex-start;
            gap: 32px;
            margin-bottom: 32px;
            page-break-inside: avoid;
        }
        .table {
            border-collapse: collapse;
            width: 300px;
            max-width: 300px;
            background: #fff;
            box-shadow: 0 1px 4px #bbb;
            page-break-inside: avoid;
        }
        th, td {
            border: 1px solid #888;
            padding: 2px 6px;
            min-width: 70px;
            font-size: 8px;
        }
        th {
            background: #e0e0e0;
            font-weight: bold;
            text-align: center;
            color: #222;
            letter-spacing: 0.5px;
        }
        td { height: 12px; }
        .firma { min-width: 50px; }
        .firma-placeholder {
            color: #aaa;
            font-style: italic;
            font-size: 8px;
        }
        /* Filas alternas en gris claro */
        tr:nth-child(even) td {
            background: #f6f6f6;
        }
        tr:nth-child(odd) td {
            background: #fff;
        }
        </style>
    </head>
    <body>
        <div class="header-container">
            <h2>Hoja de Firmas</h2>
            <div class="subject-box">
                {% if subject %}{{ subject }}{% else %}<span class="subject-placeholder">Asignatura</span>{% endif %}
            </div>
            <div class="group-box">
                {% if group %}{{ group }}{% else %}<span class="group-placeholder">Grupo</span>{% endif %}
            </div>
            <div class="date-box">
                <span class="date-placeholder">Fecha</span>
            </div>
            <div class="week-box">
                {% if week %}Semana {{ week }}{% else %}<span class="week-placeholder">Semana</span>{% endif %}
            </div>
        </div>
        {% for pair in block_pairs %}
        <div class="row">
            {% for block in pair %}
                <table class="table">
                    <tr>
                        <th>Nombre</th>
                        <th>Apellidos</th>
                        <th class="firma" colspan="2">Firma</th>
                    </tr>
                    {% for participant in block %}
                    <tr>
                        <td>{{ participant[0] }}</td>
                        <td>{{ participant[1] }}</td>
                        <td><span class="firma-placeholder">{{ day1 }}</span></td>
                        <td><span class="firma-placeholder">{{ day2 }}</span></td>
                    </tr>
                    {% endfor %}
                </table>
            {% endfor %}
        </div>
        {% endfor %}
    </body>
    </html>
    """
    template = Template(html_template)
    html_content = template.render(block_pairs=block_pairs, subject=subject, group=group, week=week, day1=day1, day2=day2)
    try:
        HTML(string=html_content).write_pdf(str(output_path))
        print(f"✓ PDF generado exitosamente en: {output_path}")
        print(f"✓ Total de participantes procesados: {len(participant_list)}")
    except Exception as e:
        print(f"Error al generar el PDF: {e}")

def process_signatures_file() -> None:
    """
    Función principal: obtiene participantes y crea hoja de firmas.
    """
    result_dir = Path("./result")
    output_path = result_dir / "hoja_firmas.xlsx"
    pdf_output_path = result_dir / "hoja_firmas.pdf"
    result_dir.mkdir(exist_ok=True)
    participant_list = get_participants()
    if not participant_list:
        return
    create_sign_sheet(participant_list, output_path)
    # Solicitar nombre de asignatura al usuario
    import sys
    subject = sys.argv[1] if len(sys.argv) > 1 and sys.argv[1].strip() else None
    group = sys.argv[2] if len(sys.argv) > 2 and sys.argv[2].strip() else None
    week = int(sys.argv[3]) if len(sys.argv) > 3 and sys.argv[3].strip().isdigit() and 1 <= int(sys.argv[3]) <= 15 else None
    # Solicitar días por teclado
    try:
        day1 = input("Introduce el nombre del primer día (dejar vacío para 'primera'): ").strip()
    except Exception:
        day1 = ''
    if not day1:
        day1 = 'primera'
    try:
        day2 = input("Introduce el nombre del segundo día (dejar vacío para 'segunda'): ").strip()
    except Exception:
        day2 = ''
    if not day2:
        day2 = 'segunda'
    create_sign_pdf(participant_list, pdf_output_path, subject, group, week, day1, day2)
    print("✓ PDF de firmas también generado en result/hoja_firmas.pdf")

if __name__ == "__main__":
    process_signatures_file()
